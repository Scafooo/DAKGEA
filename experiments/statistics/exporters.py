#!/usr/bin/env python3
"""Export utilities for statistics - CSV, Markdown, LaTeX, Excel."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List

def ensure_dir(path: Path) -> None:
    """Create directory if it doesn't exist."""
    path.mkdir(parents=True, exist_ok=True)


def write_csv(
    path: Path,
    headers: List[str],
    rows: List[List[str]],
) -> None:
    """Write CSV file with proper quoting."""
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        writer.writerows(rows)


def write_dataset_summary_csv(
    path: Path,
    dataset_stage_stats: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    metrics: List[str],
) -> None:
    """Export dataset summary as CSV."""
    headers = [
        "dataset",
        "metric",
        "reduction_mean",
        "reduction_std",
        "reduction_min",
        "reduction_max",
        "reduction_count",
        "augmentation_mean",
        "augmentation_std",
        "augmentation_min",
        "augmentation_max",
        "augmentation_count",
        "delta_mean",
        "delta_percentage",
    ]

    rows = []
    for dataset, stage_stats in sorted(dataset_stage_stats.items()):
        for metric in metrics:
            red_stats = stage_stats.get("reduction", {}).get(metric)
            aug_stats = stage_stats.get("augmentation", {}).get(metric)
            if not red_stats and not aug_stats:
                continue

            delta = None
            delta_pct = None
            if red_stats and aug_stats:
                delta = aug_stats["mean"] - red_stats["mean"]
                if red_stats["mean"] != 0:
                    delta_pct = (delta / red_stats["mean"]) * 100

            row = [
                dataset,
                metric,
                f"{red_stats['mean']:.6f}" if red_stats else "",
                f"{red_stats['std']:.6f}" if red_stats else "",
                f"{red_stats['min']:.6f}" if red_stats else "",
                f"{red_stats['max']:.6f}" if red_stats else "",
                str(red_stats["count"]) if red_stats else "",
                f"{aug_stats['mean']:.6f}" if aug_stats else "",
                f"{aug_stats['std']:.6f}" if aug_stats else "",
                f"{aug_stats['min']:.6f}" if aug_stats else "",
                f"{aug_stats['max']:.6f}" if aug_stats else "",
                str(aug_stats["count"]) if aug_stats else "",
                f"{delta:.6f}" if delta is not None else "",
                f"{delta_pct:.2f}" if delta_pct is not None else "",
            ]
            rows.append(row)

    write_csv(path, headers, rows)


def write_markdown_table(
    path: Path,
    headers: List[str],
    rows: List[List[str]],
    alignments: List[str] | None = None,
) -> None:
    """Write Markdown table."""
    ensure_dir(path.parent)

    if alignments is None:
        alignments = ["left"] * len(headers)

    with path.open("w", encoding="utf-8") as f:
        # Header
        f.write("| " + " | ".join(headers) + " |\n")

        # Separator with alignment
        sep_parts = []
        for align in alignments:
            if align == "left":
                sep_parts.append(":---")
            elif align == "right":
                sep_parts.append("---:")
            elif align == "center":
                sep_parts.append(":---:")
            else:
                sep_parts.append("---")
        f.write("| " + " | ".join(sep_parts) + " |\n")

        # Data rows
        for row in rows:
            f.write("| " + " | ".join(str(cell) for cell in row) + " |\n")


def write_dataset_summary_markdown(
    path: Path,
    dataset_stage_stats: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    metrics: List[str],
) -> None:
    """Export dataset summary as Markdown table."""
    headers = [
        "Dataset",
        "Metric",
        "Red Mean",
        "Red Std",
        "Aug Mean",
        "Aug Std",
        "Delta",
        "Δ%",
    ]

    alignments = ["left", "left", "right", "right", "right", "right", "right", "right"]

    rows = []
    for dataset, stage_stats in sorted(dataset_stage_stats.items()):
        for metric in metrics:
            red_stats = stage_stats.get("reduction", {}).get(metric)
            aug_stats = stage_stats.get("augmentation", {}).get(metric)
            if not red_stats and not aug_stats:
                continue

            delta = None
            delta_pct = None
            if red_stats and aug_stats:
                delta = aug_stats["mean"] - red_stats["mean"]
                if red_stats["mean"] != 0:
                    delta_pct = (delta / red_stats["mean"]) * 100

            row = [
                dataset,
                metric,
                f"{red_stats['mean']:.4f}" if red_stats else "—",
                f"{red_stats['std']:.4f}" if red_stats else "—",
                f"{aug_stats['mean']:.4f}" if aug_stats else "—",
                f"{aug_stats['std']:.4f}" if aug_stats else "—",
                f"+{delta:.4f}" if delta and delta > 0 else f"{delta:.4f}" if delta else "—",
                f"+{delta_pct:.1f}%" if delta_pct and delta_pct > 0 else f"{delta_pct:.1f}%" if delta_pct else "—",
            ]
            rows.append(row)

    write_markdown_table(path, headers, rows, alignments)


def escape_latex(text: str) -> str:
    """Escape special LaTeX characters."""
    replacements = {
        "_": "\\_",
        "%": "\\%",
        "$": "\\$",
        "&": "\\&",
        "#": "\\#",
        "{": "\\{",
        "}": "\\}",
        "~": "\\textasciitilde{}",
        "^": "\\textasciicircum{}",
    }
    result = str(text)
    for char, escaped in replacements.items():
        result = result.replace(char, escaped)
    return result


def write_latex_table(
    path: Path,
    headers: List[str],
    rows: List[List[str]],
    caption: str = "",
    label: str = "",
    col_spec: str | None = None,
    position: str = "htbp",
    small: bool = False,
) -> None:
    """Write LaTeX table with booktabs style.

    Args:
        path: Output file path
        headers: Column headers
        rows: Data rows
        caption: Table caption
        label: Table label for referencing
        col_spec: Column specification (e.g., "lrrr"). If None, auto-generates (left + right-aligned)
        position: Table float position (default: htbp)
        small: Use small font size
    """
    ensure_dir(path.parent)

    n_cols = len(headers)
    if col_spec is None:
        col_spec = "l" + "r" * (n_cols - 1)  # First column left, rest right-aligned

    with path.open("w", encoding="utf-8") as f:
        f.write(f"\\begin{{table}}[{position}]\n")
        f.write("\\centering\n")
        if small:
            f.write("\\small\n")
        if caption:
            f.write(f"\\caption{{{escape_latex(caption)}}}\n")
        if label:
            f.write(f"\\label{{{label}}}\n")
        f.write(f"\\begin{{tabular}}{{{col_spec}}}\n")
        f.write("\\toprule\n")

        # Header
        escaped_headers = [escape_latex(h) for h in headers]
        f.write(" & ".join(escaped_headers) + " \\\\\n")
        f.write("\\midrule\n")

        # Data rows
        for row in rows:
            escaped_row = [escape_latex(cell) for cell in row]
            f.write(" & ".join(escaped_row) + " \\\\\n")

        f.write("\\bottomrule\n")
        f.write("\\end{tabular}\n")
        f.write("\\end{table}\n")


def write_latex_table_colored(
    path: Path,
    headers: List[str],
    rows: List[List[str]],
    highlight_cols: List[int] | None = None,
    caption: str = "",
    label: str = "",
    col_spec: str | None = None,
    position: str = "htbp",
    small: bool = False,
) -> None:
    """Write LaTeX table with conditional coloring for delta columns.

    Args:
        path: Output file path
        headers: Column headers
        rows: Data rows - each row can be str or tuple (value, color_type) where color_type in ['positive', 'negative', 'neutral']
        highlight_cols: Column indices to apply conditional coloring (None = auto-detect delta columns)
        caption: Table caption
        label: Table label
        col_spec: Column specification
        position: Table float position
        small: Use small font size
    """
    ensure_dir(path.parent)

    n_cols = len(headers)
    if col_spec is None:
        col_spec = "l" + "r" * (n_cols - 1)

    # Auto-detect delta columns if not specified
    if highlight_cols is None:
        highlight_cols = [i for i, h in enumerate(headers) if "delta" in h.lower() or "Δ" in h]

    with path.open("w", encoding="utf-8") as f:
        # Write table header with xcolor package requirement
        f.write(f"% Requires: \\usepackage{{xcolor}}\n")
        f.write(f"\\begin{{table}}[{position}]\n")
        f.write("\\centering\n")
        if small:
            f.write("\\small\n")
        if caption:
            f.write(f"\\caption{{{escape_latex(caption)}}}\n")
        if label:
            f.write(f"\\label{{{label}}}\n")
        f.write(f"\\begin{{tabular}}{{{col_spec}}}\n")
        f.write("\\toprule\n")

        # Header
        escaped_headers = [escape_latex(h) for h in headers]
        f.write(" & ".join(escaped_headers) + " \\\\\n")
        f.write("\\midrule\n")

        # Data rows
        for row in rows:
            formatted_cells = []
            for col_idx, cell in enumerate(row):
                cell_str = escape_latex(str(cell))

                # Apply coloring to delta columns if cell contains +/- and is numeric
                if col_idx in highlight_cols and cell_str and cell_str not in ["—", ""]:
                    try:
                        # Extract numeric value
                        value_str = cell_str.replace("+", "").replace("\\%", "").strip()
                        value = float(value_str)
                        if value > 0:
                            cell_str = f"\\textcolor{{green!70!black}}{{{cell_str}}}"
                        elif value < 0:
                            cell_str = f"\\textcolor{{red!70!black}}{{{cell_str}}}"
                    except (ValueError, AttributeError):
                        pass

                formatted_cells.append(cell_str)

            f.write(" & ".join(formatted_cells) + " \\\\\n")

        f.write("\\bottomrule\n")
        f.write("\\end{tabular}\n")
        f.write("\\end{table}\n")


def write_dataset_summary_latex(
    path: Path,
    dataset_stage_stats: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    metrics: List[str],
    colored: bool = True,
) -> None:
    """Export dataset summary as LaTeX table with optional conditional coloring.

    Args:
        path: Output file path
        dataset_stage_stats: Nested dict with dataset -> stage -> metric -> statistics
        metrics: List of metrics to include
        colored: Apply conditional coloring to delta values
    """
    headers = [
        "Dataset",
        "Metric",
        "Red Mean",
        "Red Std",
        "Aug Mean",
        "Aug Std",
        "Delta",
        "$\\Delta$\\%",
    ]

    rows = []
    for dataset, stage_stats in sorted(dataset_stage_stats.items()):
        for metric in metrics:
            red_stats = stage_stats.get("reduction", {}).get(metric)
            aug_stats = stage_stats.get("augmentation", {}).get(metric)
            if not red_stats and not aug_stats:
                continue

            delta = None
            delta_pct = None
            if red_stats and aug_stats:
                delta = aug_stats["mean"] - red_stats["mean"]
                if red_stats["mean"] != 0:
                    delta_pct = (delta / red_stats["mean"]) * 100

            row = [
                dataset,
                metric,
                f"{red_stats['mean']:.4f}" if red_stats else "—",
                f"[[PM]]{red_stats['std']:.4f}" if red_stats else "—",
                f"{aug_stats['mean']:.4f}" if aug_stats else "—",
                f"[[PM]]{aug_stats['std']:.4f}" if aug_stats else "—",
                f"+{delta:.4f}" if delta and delta > 0 else f"{delta:.4f}" if delta else "—",
                f"+{delta_pct:.2f}%" if delta_pct and delta_pct > 0 else f"{delta_pct:.2f}%" if delta_pct else "—",
            ]
            rows.append(row)

    if colored:
        write_latex_table_colored(
            path,
            headers,
            rows,
            caption="Reduction vs Augmentation Performance Comparison",
            label="tab:reduction_augmentation_comparison",
            small=True,
        )
    else:
        write_latex_table(
            path,
            headers,
            rows,
            caption="Reduction vs Augmentation Performance Comparison",
            label="tab:reduction_augmentation_comparison",
            small=True,
        )


def try_write_excel(
    path: Path,
    dataset_stage_stats: Dict[str, Dict[str, Dict[str, Dict[str, float]]]],
    metrics: List[str],
) -> bool:
    """Try to export to Excel if openpyxl is available."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False

    ensure_dir(path.parent)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Headers
    headers = [
        "Dataset",
        "Metric",
        "Red Mean",
        "Red Std",
        "Red Min",
        "Red Max",
        "Red N",
        "Aug Mean",
        "Aug Std",
        "Aug Min",
        "Aug Max",
        "Aug N",
        "Delta",
        "Delta %",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data
    row_idx = 2
    for dataset, stage_stats in sorted(dataset_stage_stats.items()):
        for metric in metrics:
            red_stats = stage_stats.get("reduction", {}).get(metric)
            aug_stats = stage_stats.get("augmentation", {}).get(metric)
            if not red_stats and not aug_stats:
                continue

            delta = None
            delta_pct = None
            if red_stats and aug_stats:
                delta = aug_stats["mean"] - red_stats["mean"]
                if red_stats["mean"] != 0:
                    delta_pct = (delta / red_stats["mean"]) * 100

            ws.cell(row=row_idx, column=1, value=dataset)
            ws.cell(row=row_idx, column=2, value=metric)
            ws.cell(row=row_idx, column=3, value=red_stats["mean"] if red_stats else None)
            ws.cell(row=row_idx, column=4, value=red_stats["std"] if red_stats else None)
            ws.cell(row=row_idx, column=5, value=red_stats["min"] if red_stats else None)
            ws.cell(row=row_idx, column=6, value=red_stats["max"] if red_stats else None)
            ws.cell(row=row_idx, column=7, value=red_stats["count"] if red_stats else None)
            ws.cell(row=row_idx, column=8, value=aug_stats["mean"] if aug_stats else None)
            ws.cell(row=row_idx, column=9, value=aug_stats["std"] if aug_stats else None)
            ws.cell(row=row_idx, column=10, value=aug_stats["min"] if aug_stats else None)
            ws.cell(row=row_idx, column=11, value=aug_stats["max"] if aug_stats else None)
            ws.cell(row=row_idx, column=12, value=aug_stats["count"] if aug_stats else None)
            ws.cell(row=row_idx, column=13, value=delta)
            ws.cell(row=row_idx, column=14, value=delta_pct)

            # Highlight positive deltas in green
            if delta and delta > 0:
                ws.cell(row=row_idx, column=13).fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )

            row_idx += 1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(path)
    return True


def write_comparison_tables_latex(
    output_dir: Path,
    ratio_entries: Dict[str, List[Dict]],
    metrics: List[str] | None = None,
) -> int:
    """Generate LaTeX comparison tables (one per dataset) with ratio-based comparisons.

    Creates tables showing baseline vs augmented results side-by-side for each metric,
    with rows representing reduction ratios.

    Args:
        output_dir: Directory to write .tex files
        ratio_entries: Dict[dataset -> List[experiment entries]]
                       Each entry has 'reduction' and 'augmentation' (or 'plm') with metrics
        metrics: List of metrics to include. If None, uses default set.

    Returns:
        Number of tables generated
    """
    from statistics import mean, pstdev

    if metrics is None:
        metrics = [
            ('hits@1', 'H@1', True, 2, True),
            ('hits@5', 'H@5', True, 2, True),
            ('hits@10', 'H@10', True, 2, True),
            ('mrr', 'MRR', False, 4, True),
            ('mr', 'MR', False, 1, False),
            ('precision', 'P', True, 2, True),
            ('recall', 'R', True, 2, True),
            ('f-measure', 'F1', True, 2, True),
        ]
    else:
        # Convert simple metric names to tuple format
        metric_configs = {
            'hits@1': ('hits@1', 'H@1', True, 2, True),
            'hits@5': ('hits@5', 'H@5', True, 2, True),
            'hits@10': ('hits@10', 'H@10', True, 2, True),
            'mrr': ('mrr', 'MRR', False, 4, True),
            'mr': ('mr', 'MR', False, 1, False),
            'precision': ('precision', 'P', True, 2, True),
            'recall': ('recall', 'R', True, 2, True),
            'f-measure': ('f-measure', 'F1', True, 2, True),
        }
        metrics = [metric_configs.get(m, (m, m.upper(), False, 3, True)) for m in metrics]

    ensure_dir(output_dir)

    # Aggregate by dataset and ratio
    aggregated = {}
    for dataset, entries in ratio_entries.items():
        ratios_data = {}
        for entry in entries:
            red_data = entry.get("reduction")
            aug_data = entry.get("augmentation") or entry.get("plm")

            if not red_data or not aug_data:
                continue

            # Get reduction ratio
            try:
                red_ratio = round(float(red_data.get("ratio")), 6)
            except (TypeError, ValueError):
                continue

            # Initialize ratio entry
            if red_ratio not in ratios_data:
                ratios_data[red_ratio] = {
                    'baseline': {m[0]: [] for m in metrics},
                    'augmented': {m[0]: [] for m in metrics}
                }

            # Collect metrics
            red_metrics = red_data.get("metrics", {})
            aug_metrics = aug_data.get("metrics", {})

            for metric_key, _, _, _, _ in metrics:
                if metric_key in red_metrics:
                    ratios_data[red_ratio]['baseline'][metric_key].append(red_metrics[metric_key])
                if metric_key in aug_metrics:
                    ratios_data[red_ratio]['augmented'][metric_key].append(aug_metrics[metric_key])

        if ratios_data:
            aggregated[dataset] = ratios_data

    # Generate tables
    tables_generated = 0
    # All possible reduction ratios from 0.1 to 1.0
    all_ratios = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]

    for dataset, ratios_data in sorted(aggregated.items()):
        output_file = output_dir / f"{dataset}.tex"

        # Build LaTeX table
        latex = []
        latex.append(r"% Required packages: \usepackage{multirow,xcolor,colortbl,graphicx,float}")
        latex.append(r"% Values shown as: mean±std (standard deviation)")
        latex.append(r"\begin{table}[H]")
        latex.append(r"\centering")
        latex.append(r"\scriptsize")
        latex.append(r"\caption{Results for " + escape_latex(dataset) + r". Values shown as mean$\pm$std across multiple seeds.}")
        latex.append(r"\label{tab:" + dataset.lower() + r"}")

        # Add resizebox to fit table to text width (slightly wider to improve readability)
        latex.append(r"\resizebox{1.1\textwidth}{!}{%")

        # Column spec: Ratio | pairs of (Base/Aug) for each metric
        num_metrics = len(metrics)
        col_spec = "c|" + "cc|" * num_metrics
        latex.append(r"\begin{tabular}{" + col_spec + r"}")
        latex.append(r"\hline")

        # Header row 1: Metric names spanning 2 columns
        header1 = r"\multirow{2}{*}{\textbf{Ratio}}"
        for _, metric_name, _, _, _ in metrics:
            header1 += f" & \\multicolumn{{2}}{{c|}}{{{escape_latex(metric_name)}}}"
        header1 += r" \\"
        latex.append(header1)

        # Header row 2: Base/Aug for each metric
        header2 = " "
        for _ in metrics:
            header2 += r" & \textit{Base} & \textit{Aug}"
        header2 += r" \\"
        latex.append(header2)
        latex.append(r"\hline")

        # Data rows - always show all ratios from 0.1 to 1.0
        for ratio in all_ratios:
            data = ratios_data.get(ratio)  # May be None if no data for this ratio
            row = f"\\textbf{{{ratio:.1f}}}"

            for metric_key, _, is_percentage, decimals, higher_is_better in metrics:
                # If no data for this ratio, show N/A for both baseline and augmented
                if data is None:
                    row += r" & \textcolor{gray}{N/A}"  # Baseline
                    row += r" & \textcolor{gray}{N/A}"  # Augmented
                    continue

                # Compute statistics
                baseline_values = data['baseline'].get(metric_key, [])
                augmented_values = data['augmented'].get(metric_key, [])

                baseline_mean = mean(baseline_values) if baseline_values else None
                baseline_std = pstdev(baseline_values) if len(baseline_values) > 1 else (0.0 if baseline_values else None)

                augmented_mean = mean(augmented_values) if augmented_values else None
                augmented_std = pstdev(augmented_values) if len(augmented_values) > 1 else (0.0 if augmented_values else None)

                # Determine which value is better (for bold formatting)
                baseline_is_better = False
                augmented_is_better = False

                if baseline_mean is not None and augmented_mean is not None:
                    if higher_is_better:
                        if baseline_mean > augmented_mean:
                            baseline_is_better = True
                        else:
                            augmented_is_better = True
                    else:  # Lower is better (e.g., MR)
                        if baseline_mean < augmented_mean:
                            baseline_is_better = True
                        else:
                            augmented_is_better = True

                # Baseline column
                if baseline_mean is None:
                    row += r" & \textcolor{gray}{N/A}"
                else:
                    if is_percentage:
                        baseline_mean *= 100
                        baseline_std *= 100
                    # Bold if this is the better value
                    if baseline_is_better:
                        row += f" & \\textbf{{{baseline_mean:.{decimals}f}$\\pm${baseline_std:.{decimals}f}}}"
                    else:
                        row += f" & {baseline_mean:.{decimals}f}$\\pm${baseline_std:.{decimals}f}"

                # Augmented column with color
                if augmented_mean is None:
                    row += r" & \textcolor{gray}{N/A}"
                else:
                    if is_percentage:
                        augmented_mean *= 100
                        augmented_std *= 100

                    # Determine color
                    color_cmd = ""
                    if baseline_mean is not None:
                        if higher_is_better:
                            is_improvement = augmented_mean > baseline_mean
                        else:
                            is_improvement = augmented_mean < baseline_mean

                        if is_improvement:
                            color_cmd = r"\cellcolor{green!15}"
                        else:
                            color_cmd = r"\cellcolor{red!15}"

                    # Bold if this is the better value
                    if augmented_is_better:
                        row += f" & {color_cmd}\\textbf{{{augmented_mean:.{decimals}f}$\\pm${augmented_std:.{decimals}f}}}"
                    else:
                        row += f" & {color_cmd}{augmented_mean:.{decimals}f}$\\pm${augmented_std:.{decimals}f}"

            row += r" \\"
            latex.append(row)

        latex.append(r"\hline")
        latex.append(r"\end{tabular}")
        latex.append(r"}% End resizebox")
        latex.append(r"\end{table}")

        # Write file
        with open(output_file, 'w') as f:
            f.write('\n'.join(latex))

        tables_generated += 1

    return tables_generated


def write_detailed_comparison_tables_latex(
    output_dir: Path,
    ratio_entries: Dict[str, List[Dict]],
    metrics: List[str] | None = None,
) -> int:
    """Generate detailed LaTeX tables showing individual experiment values (not aggregated).

    Creates one table per dataset per reduction ratio, showing all individual seeds
    with their baseline vs augmented values side-by-side.

    Args:
        output_dir: Directory to write .tex files
        ratio_entries: Dict[dataset -> List[experiment entries]]
        metrics: List of metrics to include. If None, uses default set.

    Returns:
        Number of tables generated
    """
    from statistics import mean, pstdev

    if metrics is None:
        metrics = [
            ('hits@1', 'H@1', True, 2, True),
            ('hits@5', 'H@5', True, 2, True),
            ('hits@10', 'H@10', True, 2, True),
            ('mrr', 'MRR', False, 4, True),
            ('mr', 'MR', False, 1, False),
            ('precision', 'P', True, 2, True),
            ('recall', 'R', True, 2, True),
            ('f-measure', 'F1', True, 2, True),
        ]
    else:
        # Convert simple metric names to tuple format
        metric_configs = {
            'hits@1': ('hits@1', 'H@1', True, 2, True),
            'hits@5': ('hits@5', 'H@5', True, 2, True),
            'hits@10': ('hits@10', 'H@10', True, 2, True),
            'mrr': ('mrr', 'MRR', False, 4, True),
            'mr': ('mr', 'MR', False, 1, False),
            'precision': ('precision', 'P', True, 2, True),
            'recall': ('recall', 'R', True, 2, True),
            'f-measure': ('f-measure', 'F1', True, 2, True),
        }
        metrics = [metric_configs.get(m, (m, m.upper(), False, 3, True)) for m in metrics]

    ensure_dir(output_dir)

    # Organize by dataset and reduction ratio
    organized = {}
    for dataset, entries in ratio_entries.items():
        dataset_ratios = {}
        for entry in entries:
            red_data = entry.get("reduction")
            aug_data = entry.get("augmentation") or entry.get("plm")

            if not red_data or not aug_data:
                continue

            # Get reduction ratio and experiment name
            try:
                red_ratio = round(float(red_data.get("ratio")), 6)
            except (TypeError, ValueError):
                continue

            exp_name = entry.get("experiment", "unknown")

            # Initialize ratio entry
            if red_ratio not in dataset_ratios:
                dataset_ratios[red_ratio] = []

            # Store individual experiment data
            dataset_ratios[red_ratio].append({
                'experiment': exp_name,
                'baseline_metrics': red_data.get("metrics", {}),
                'augmented_metrics': aug_data.get("metrics", {}),
            })

        if dataset_ratios:
            organized[dataset] = dataset_ratios

    # Generate tables
    tables_generated = 0
    # All possible reduction ratios from 0.1 to 1.0
    all_ratios = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]

    for dataset, ratios_data in sorted(organized.items()):
        for ratio in all_ratios:  # Generate table for ALL ratios 0.1-1.0
            experiments = ratios_data.get(ratio, [])  # Get experiments or empty list

            # Sort experiments by name if present
            if experiments:
                experiments = sorted(experiments, key=lambda x: x['experiment'])

            output_file = output_dir / f"{dataset}_ratio{ratio:.1f}_detailed.tex"

            # Build LaTeX table
            latex = []
            latex.append(r"% Required packages: \usepackage{multirow,xcolor,colortbl,graphicx,float}")
            latex.append(r"% Individual experiment values (not aggregated)")
            latex.append(r"\begin{table}[H]")
            latex.append(r"\centering")
            latex.append(r"\scriptsize")
            latex.append(r"\caption{Detailed results for " + escape_latex(dataset) +
                        f" at reduction ratio {ratio:.1f}. Individual experiment values.}}")
            latex.append(r"\label{tab:" + dataset.lower() + f"_ratio{ratio:.1f}_detailed" + r"}")

            # Add resizebox to fit table to text width
            latex.append(r"\resizebox{\textwidth}{!}{%")

            # Column spec: Experiment | triplets of (Base/Aug/Gap) for each metric
            num_metrics = len(metrics)
            col_spec = "l|" + "ccc|" * num_metrics
            latex.append(r"\begin{tabular}{" + col_spec + r"}")
            latex.append(r"\hline")

            # Header row 1: Metric names spanning 3 columns (Base, Aug, Gap)
            header1 = r"\multirow{2}{*}{\textbf{Experiment}}"
            for _, metric_name, _, _, _ in metrics:
                header1 += f" & \\multicolumn{{3}}{{c|}}{{{escape_latex(metric_name)}}}"
            header1 += r" \\"
            latex.append(header1)

            # Header row 2: Base/Aug/Δ for each metric
            header2 = " "
            for _ in metrics:
                header2 += r" & \textit{Base} & \textit{Aug} & \textit{$\Delta$}"
            header2 += r" \\"
            latex.append(header2)
            latex.append(r"\hline")

            # Data rows - one per experiment
            if not experiments:
                # No data for this ratio - show empty row
                row = r"\textit{No experiments for this ratio}"
                for _ in metrics:
                    row += r" & \textcolor{gray}{N/A} & \textcolor{gray}{N/A} & \textcolor{gray}{N/A}"
                row += r" \\"
                latex.append(row)
            else:
                for exp_data in experiments:
                    exp_name = exp_data['experiment']
                    baseline_metrics = exp_data['baseline_metrics']
                    augmented_metrics = exp_data['augmented_metrics']

                    row = f"\\texttt{{{escape_latex(exp_name)}}}"

                    for metric_key, _, is_percentage, decimals, higher_is_better in metrics:
                        # Get individual values
                        baseline_value = baseline_metrics.get(metric_key)
                        augmented_value = augmented_metrics.get(metric_key)

                        # Determine which value is better (for bold formatting)
                        baseline_is_better = False
                        augmented_is_better = False

                        if baseline_value is not None and augmented_value is not None:
                            if higher_is_better:
                                if baseline_value > augmented_value:
                                    baseline_is_better = True
                                else:
                                    augmented_is_better = True
                            else:  # Lower is better (e.g., MR)
                                if baseline_value < augmented_value:
                                    baseline_is_better = True
                                else:
                                    augmented_is_better = True

                        # Calculate gap (before percentage conversion for correct calculation)
                        gap = None
                        if baseline_value is not None and augmented_value is not None:
                            gap = augmented_value - baseline_value

                        # Baseline column
                        if baseline_value is None:
                            row += r" & \textcolor{gray}{N/A}"
                        else:
                            if is_percentage:
                                baseline_value *= 100
                            # Bold if this is the better value
                            if baseline_is_better:
                                row += f" & \\textbf{{{baseline_value:.{decimals}f}}}"
                            else:
                                row += f" & {baseline_value:.{decimals}f}"

                        # Augmented column with color
                        if augmented_value is None:
                            row += r" & \textcolor{gray}{N/A}"
                        else:
                            if is_percentage:
                                augmented_value *= 100

                            # Determine color
                            color_cmd = ""
                            if baseline_value is not None:
                                if higher_is_better:
                                    is_improvement = augmented_value > baseline_value
                                else:
                                    is_improvement = augmented_value < baseline_value

                                if is_improvement:
                                    color_cmd = r"\cellcolor{green!15}"
                                else:
                                    color_cmd = r"\cellcolor{red!15}"

                            # Bold if this is the better value
                            if augmented_is_better:
                                row += f" & {color_cmd}\\textbf{{{augmented_value:.{decimals}f}}}"
                            else:
                                row += f" & {color_cmd}{augmented_value:.{decimals}f}"

                        # Gap column (Δ)
                        if gap is None:
                            row += r" & \textcolor{gray}{N/A}"
                        else:
                            if is_percentage:
                                gap *= 100
                            # Color based on whether it's an improvement
                            if higher_is_better:
                                gap_color = "green!70!black" if gap > 0 else "red!70!black"
                            else:  # Lower is better (e.g., MR)
                                gap_color = "green!70!black" if gap < 0 else "red!70!black"

                            # Format with sign
                            gap_str = f"{gap:+.{decimals}f}"
                            row += f" & \\textcolor{{{gap_color}}}{{{gap_str}}}"

                    row += r" \\"
                    latex.append(row)

            latex.append(r"\hline")
            latex.append(r"\end{tabular}")
            latex.append(r"}% End resizebox")
            latex.append(r"\end{table}")

            # Write file
            with open(output_file, 'w') as f:
                f.write('\n'.join(latex))

            tables_generated += 1

    # Generate combined file for each dataset (includes all ratios)
    for dataset in sorted(organized.keys()):
        combined_file = output_dir / f"{dataset}_all_detailed.tex"

        combined_latex = []
        combined_latex.append(r"% Combined detailed tables for " + escape_latex(dataset))
        combined_latex.append(r"% Includes all reduction ratios 0.1-1.0")
        combined_latex.append(r"% Required packages: \usepackage{multirow,xcolor,colortbl,graphicx,float}")
        combined_latex.append("")

        for ratio in all_ratios:
            # Input individual table file (use relative path since files are in same directory)
            individual_file = f"{dataset}_ratio{ratio:.1f}_detailed.tex"
            combined_latex.append(f"% Ratio {ratio:.1f}")
            combined_latex.append(f"\\input{{{individual_file}}}")
            combined_latex.append("")

        # Write combined file
        with open(combined_file, 'w') as f:
            f.write('\n'.join(combined_latex))

    return tables_generated


__all__ = [
    "write_dataset_summary_csv",
    "write_dataset_summary_markdown",
    "write_dataset_summary_latex",
    "write_comparison_tables_latex",
    "write_detailed_comparison_tables_latex",
    "write_latex_table",
    "write_latex_table_colored",
    "escape_latex",
    "try_write_excel",
    "ensure_dir",
]
